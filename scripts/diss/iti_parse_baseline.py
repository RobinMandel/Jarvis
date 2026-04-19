import openpyxl, statistics

path = r'E:\OneDrive\Dokumente\Studium\!!Medizin\03 Doktorarbeit\ITI\Cardiochirurgie\Klinische Daten\Kardiochirurgie_Klinische_Daten_DJ_final.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Tabelle1']
ages, sex_m, sex_w = [], 0, 0
aht, hlp, dm2 = [], [], []
ecm_dur, isch_dur, repf_dur = [], [], []
transfusion_eks = []

for r in range(2, 28):
    nr = ws.cell(r,1).value
    if not nr or not str(nr).isdigit():
        break
    age = ws.cell(r,2).value
    sex = ws.cell(r,3).value
    if age: ages.append(int(age))
    if sex == 'm': sex_m += 1
    elif sex == 'w': sex_w += 1
    aht.append(ws.cell(r,5).value)
    hlp.append(ws.cell(r,6).value)
    dm2.append(ws.cell(r,7).value)
    ecm = ws.cell(r,28).value
    isch = ws.cell(r,29).value
    rep = ws.cell(r,30).value
    eks = ws.cell(r,33).value
    if ecm: ecm_dur.append(str(ecm))
    if isch: isch_dur.append(str(isch))
    if rep: repf_dur.append(str(rep))
    if eks is not None: transfusion_eks.append(eks)

print(f"n={len(ages)} Patienten (KC_01-KC_26)")
print(f"Alter: Median={statistics.median(ages):.0f}, Mean={statistics.mean(ages):.1f}, Range={min(ages)}-{max(ages)}")
print(f"Geschlecht: {sex_m}m / {sex_w}w ({100*sex_m//(sex_m+sex_w)}% maennlich)")
print(f"aHT: {aht.count('ja')}/{len(aht)}")
print(f"HLP: {hlp.count('ja')}/{len(hlp)}")
print(f"DM2: {dm2.count('ja')}/{len(dm2)}")
print(f"ECM-Dauer sample: {ecm_dur[:4]}")
print(f"Ischaemie-Dauer sample: {isch_dur[:4]}")
print(f"EKs sample: {transfusion_eks[:8]}")
wb.close()

# Now KDIGO
kdigo_path = r'E:\OneDrive\Dokumente\Studium\!!Medizin\03 Doktorarbeit\ITI\Cardiochirurgie\Baseline TRAKI KDIGO\KDIGO 15-26.xlsx'
if __import__('os').path.exists(kdigo_path):
    wb2 = openpyxl.load_workbook(kdigo_path, data_only=True)
    print("\n=== KDIGO 15-26.xlsx ===")
    for sh in wb2.sheetnames:
        ws2 = wb2[sh]
        print(f"  Sheet: {sh} | {ws2.max_row} rows x {ws2.max_column} cols")
        hdrs = [str(ws2.cell(1,c).value) for c in range(1, min(ws2.max_column+1, 30))]
        print(f"  Headers: {hdrs}")
        for rr in range(2, min(6, ws2.max_row+1)):
            print(f"  Row{rr}: {[str(ws2.cell(rr,c).value) for c in range(1, min(ws2.max_column+1, 15))]}")
    wb2.close()
