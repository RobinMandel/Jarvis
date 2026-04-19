import os, json, re
from openpyxl import load_workbook
from collections import Counter,defaultdict

files = [
 r'E:\OneDrive\Dokumente\Studium\!!Medizin\03 Doktorarbeit\ITI\Cardiochirurgie\Klinische Daten\Kardio_Klinische_Daten_Patienten_KC_01_bis_KC_14.xlsx',
 r'E:\OneDrive\Dokumente\Studium\!!Medizin\03 Doktorarbeit\ITI\Cardiochirurgie\Klinische Daten\Kardio_Klinische_Daten_Patienten_KC_15_bis_KC_26.xlsx',
 r'E:\OneDrive\Dokumente\Studium\!!Medizin\03 Doktorarbeit\ITI\Cardiochirurgie\Klinische Daten\20240117_OP_Technik_Darko_KC_01_bis_KC_26.xlsx',
 r'E:\OneDrive\Dokumente\Studium\!!Medizin\03 Doktorarbeit\ITI\Cardiochirurgie\Klinische Daten\FeNa.xlsx',
 r'E:\OneDrive\Dokumente\Studium\!!Medizin\03 Doktorarbeit\ITI\Cardiochirurgie\Klinische Daten\Kardio_und_Sepsis_Studie_Urin_Übersicht.xlsx',
 r'E:\OneDrive\Dokumente\Studium\!!Medizin\03 Doktorarbeit\ITI\Cardiochirurgie\Baseline TRAKI KDIGO\KDIGO 15-26.xlsx',
 r'E:\OneDrive\Dokumente\Studium\!!Medizin\03 Doktorarbeit\ITI\Cardiochirurgie\Baseline TRAKI KDIGO\Nierenparameter_Kardio_Sepsis_16.01.2023.xlsx',
]
pat = re.compile(r'KC[_ ]?\d+', re.I)
code_pat = re.compile(r'^(A|B|C|D|E|OR|24h|48h|120h|HV)$', re.I)
for f in files:
    print('\nFILE', os.path.basename(f))
    wb = load_workbook(f, read_only=True, data_only=True)
    print('SHEETS', wb.sheetnames)
    for s in wb.sheetnames[:8]:
        ws = wb[s]
        rows=[]
        for i,row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals=[v for v in row if v not in (None,'')]
            if vals:
                rows.append((i,row))
            if len(rows)>=8:
                break
        print(' SHEET', s)
        for i,row in rows[:5]:
            trimmed=[str(v)[:40] if v is not None else '' for v in row[:12]]
            print('  R',i, trimmed)
