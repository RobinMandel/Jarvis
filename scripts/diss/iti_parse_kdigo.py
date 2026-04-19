import openpyxl

kdigo_path = r'E:\OneDrive\Dokumente\Studium\!!Medizin\03 Doktorarbeit\ITI\Cardiochirurgie\Baseline TRAKI KDIGO\KDIGO 15-26.xlsx'
wb = openpyxl.load_workbook(kdigo_path, data_only=True)
ws = wb['KDIGO']

aki_patients = {}
kdigo_stages = {0:0, 1:0, 2:0, 3:0}
for r in range(5, ws.max_row+1):
    pat = ws.cell(r,1).value
    mzp = ws.cell(r,2).value
    krea_umol = ws.cell(r,8).value
    anstieg15 = ws.cell(r,10).value
    anstieg03 = ws.cell(r,11).value
    if pat and str(pat).startswith('KC'):
        has_aki = bool(anstieg15 or anstieg03)
        if has_aki:
            aki_patients[str(pat)] = str(mzp)

print("KDIGO AKI-Patienten KC_15-26:")
print(f"  Gesamt mit Kriterium: {len(aki_patients)}")
for p,m in aki_patients.items():
    print(f"  {p} bei MZP={m}")

# Neutrophilenratio lesen
nratio_path = r'E:\OneDrive\Dokumente\Studium\!!Medizin\03 Doktorarbeit\ITI\Cardiochirurgie\Neutrophilenratio.xlsx'
wb2 = openpyxl.load_workbook(nratio_path, data_only=True)
print("\nNeutrophilenratio.xlsx sheets:")
for sh in wb2.sheetnames:
    ws2 = wb2[sh]
    print(f"  {sh}: {ws2.max_row}r x {ws2.max_column}c")
    print(f"  H1: {[str(ws2.cell(1,c).value) for c in range(1, min(ws2.max_column+1,20))]}")
    print(f"  R2: {[str(ws2.cell(2,c).value) for c in range(1, min(ws2.max_column+1,15))]}")
wb2.close()
wb.close()
